# ============================================================
# Amazon/excel_writer.py
# مسؤول بس عن الحفظ (JSON + Excel). فيه شيتين في ملف الإكسيل:
#   1) "Offers"  -> جدول المقارنة (الأعمدة المطلوبة بس)
#   2) "Products" -> بيانات المنتج الأساسي (عنوان/وصف/صورة) مرة
#      واحدة لكل ASIN، مش متكررة لكل عرض
# ============================================================

import json
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --------------------------------------------------------------
# أعمدة جدول المقارنة - بالظبط الحقول المطلوبة، من غير أي زيادة
# --------------------------------------------------------------
OFFERS_HEADERS = [
    "ASIN", "Product URL", "Date", "Offer ID", "Seller",
    "Product Price", "Original Price", "Discount Amount",
    "Shipping", "Total", "Seller Rating",
]

PRODUCTS_HEADERS = [
    "ASIN", "Product Title", "Description", "Image URL", "Product URL",
]


def append_to_json_file(new_records, path):
    existing = []
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            if not isinstance(existing, list):
                existing = []
        except (json.JSONDecodeError, OSError):
            existing = []

    existing.extend(new_records)

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"[+] JSON updated: {path} (إجمالي السجلات دلوقتي: {len(existing)})")


def _style_header(ws, headers, header_fill, header_font, border):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _open_or_create_workbook(path):
    """
    لو الملف موجود بيفتحه، ولو الشيتات الأساسية ("Offers"/"Products")
    مش موجودة أو عناوينها قديمة، بياخد نسخة احتياطية ويبدأ ملف جديد
    نظيف بنفس الأعمدة المطلوبة. كده مبنفقدش بيانات قديمة، والملف
    الحالي يفضل متوافق دايمًا مع الشكل الجديد.
    """
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

    needs_fresh_file = False

    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            offers_ok = "Offers" in wb.sheetnames and \
                [c.value for c in wb["Offers"][1]] == OFFERS_HEADERS
            products_ok = "Products" in wb.sheetnames and \
                [c.value for c in wb["Products"][1]] == PRODUCTS_HEADERS
            if not (offers_ok and products_ok):
                needs_fresh_file = True
        except Exception:
            needs_fresh_file = True

        if needs_fresh_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = path.replace(".xlsx", f"_old_backup_{timestamp}.xlsx")
            try:
                os.rename(path, backup_path)
                print(f"[!] شكل الأعمدة في الملف القديم مختلف — اتحفظ نسخة احتياطية منه: {backup_path}")
                print(f"[+] هيتعمل ملف جديد نظيف: {path}")
            except OSError as e:
                print(f"[!] تعذر عمل نسخة احتياطية من الملف القديم ({e}). "
                      f"لو الملف مفتوح في Excel، اقفله وجرب تاني.")
                raise

    if os.path.exists(path):
        wb = load_workbook(path)
        offers_ws = wb["Offers"]
        products_ws = wb["Products"]
    else:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        wb = Workbook()

        offers_ws = wb.active
        offers_ws.title = "Offers"
        _style_header(offers_ws, OFFERS_HEADERS, header_fill, header_font, border)
        widths = {"A": 14, "B": 45, "C": 18, "D": 16, "E": 30,
                  "F": 14, "G": 14, "H": 15, "I": 12, "J": 12, "K": 14}
        for col, width in widths.items():
            offers_ws.column_dimensions[col].width = width
        offers_ws.freeze_panes = "A2"
        offers_ws.sheet_view.rightToLeft = True

        products_ws = wb.create_sheet("Products")
        _style_header(products_ws, PRODUCTS_HEADERS, header_fill, header_font, border)
        widths = {"A": 14, "B": 45, "C": 70, "D": 60, "E": 45}
        for col, width in widths.items():
            products_ws.column_dimensions[col].width = width
        products_ws.freeze_panes = "A2"
        products_ws.sheet_view.rightToLeft = True

    return wb, offers_ws, products_ws


def _remove_rows_for_asins(ws, asins, asin_col=1):
    """بيمسح أي صفوف قديمة تخص نفس الـ ASINs دي من الشيت (لو موجودة)،
    عشان نضمن إن الملف بيحتفظ بآخر تحديث بس من غير تكرار قديم/جديد."""
    if not asins:
        return
    rows_to_delete = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(r, asin_col).value in asins
    ]
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)


def _remove_rows_for_asins(ws, asins, asin_col=1):
    """بيمسح أي صفوف قديمة تخص نفس الـ ASINs دي من الشيت (لو موجودة)،
    عشان نضمن إن الملف بيحتفظ بآخر تحديث بس من غير تكرار قديم/جديد."""
    if not asins:
        return
    rows_to_delete = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(r, asin_col).value in asins
    ]
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)


def append_to_excel_file(new_offer_records, product_record, path, my_seller_name=None):
    """
    new_offer_records: list من صفوف عروض المقارنة (dicts) لنفس المنتج
    product_record: dict فيه بيانات المنتج الأساسي (عنوان/وصف/صورة) - يتكتب مرة واحدة بس
    ده أرشيف تاريخي (append) — بيضيف صفوف جديدة فوق القديمة، مفيد
    لتتبع تغير الأسعار مع الوقت. للنسخة "آخر تحديث فقط"، استخدم
    write_last_data_file بدل الدالة دي.
    """
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
    my_offer_fill = PatternFill("solid", fgColor="BDD7EE")
    hyperlink_font = Font(color="0563C1", underline="single")

    wb, offers_ws, products_ws = _open_or_create_workbook(path)

    # -------- صف عروض المقارنة --------
    start_row = offers_ws.max_row + 1
    for offer in new_offer_records:
        values = [
            offer.get("asin"), offer.get("product_url"), offer.get("date"),
            offer.get("offer_id"), offer.get("seller"), offer.get("product_price"),
            offer.get("original_price"), offer.get("discount_amount"),
            offer.get("shipping"), offer.get("total"), offer.get("seller_rating"),
        ]
        for col, value in enumerate(values, 1):
            cell = offers_ws.cell(start_row, col, value if value is not None else "-")
            cell.border = border
            if col in [6, 7, 8, 9, 10]:
                cell.number_format = '0.00 "SAR"'
            if col == 11 and isinstance(value, (int, float)):
                cell.number_format = '0.0'
            if col == 2 and value and value != "-":
                cell.hyperlink = value
                cell.font = hyperlink_font

        is_my_offer = (
            my_seller_name and offer.get("seller")
            and my_seller_name.strip().lower() in offer["seller"].strip().lower()
        )
        if is_my_offer:
            for col in range(1, len(OFFERS_HEADERS) + 1):
                offers_ws.cell(start_row, col).fill = my_offer_fill

        start_row += 1

    offers_ws.auto_filter.ref = f"A1:K{offers_ws.max_row}"

    # -------- صف بيانات المنتج الأساسي (مرة واحدة بس لكل ASIN) --------
    existing_asins = {
        products_ws.cell(r, 1).value
        for r in range(2, products_ws.max_row + 1)
    }
    if product_record and product_record.get("asin") not in existing_asins:
        row = products_ws.max_row + 1
        values = [
            product_record.get("asin"),
            product_record.get("title"),
            product_record.get("description"),
            product_record.get("image_url"),
            product_record.get("product_url"),
        ]
        for col, value in enumerate(values, 1):
            cell = products_ws.cell(row, col, value if value is not None else "-")
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [3]))
            if col in [4, 5] and value and value != "-":
                cell.hyperlink = value
                cell.font = hyperlink_font
        products_ws.auto_filter.ref = f"A1:E{products_ws.max_row}"

    try:
        wb.save(path)
        print(f"[+] Excel updated: {path} (إجمالي صفوف العروض دلوقتي: {offers_ws.max_row - 1})")
    except PermissionError:
        print(f"[!] الملف {path} مفتوح في برنامج تاني (زي Excel) — اقفله وجرب تاني.")
        raise


# ================================================================
# "آخر تحديث فقط" — بيعيد بناء الملف بالكامل في الذاكرة كل مرة (مش
# delete_rows على openpyxl، لأن المكتبة معروف عنها إنها بتسيب صفوف
# فاضية/تالفة أحيانًا بعد الحذف). النتيجة: سجل واحد بس حديث لكل ASIN.
# ================================================================

def _read_existing_rows(path, sheet_name):
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            return []
        return [row for row in wb[sheet_name].iter_rows(min_row=2, values_only=True) if row and row[0]]
    except Exception:
        return []


def write_last_data_file(new_offer_records, product_record, path, my_seller_name=None):
    """
    new_offer_records / product_record: نفس شكل append_to_excel_file.
    بيحتفظ بسجل واحد بس حديث لكل ASIN — أي بيانات قديمة لنفس الـ ASIN
    بتتستبدل بالكامل بالبيانات الجديدة، والمنتجات التانية اللي مش في
    الدفعة دي بتفضل زي ما هي.
    """
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    my_offer_fill = PatternFill("solid", fgColor="BDD7EE")
    hyperlink_font = Font(color="0563C1", underline="single")

    # مهم: بنفصل مجموعة الـ ASINs المستخدمة لتصفية كل شيت عن التانية.
    # لو استخدمنا نفس المجموعة الموحّدة للاتنين، أي رن (rescan) بيرجع
    # منتج من غير عروض (offers فاضية) أو من غير بيانات منتج (title/desc
    # فاضيين) كان بيمسح الصفوف القديمة الصحيحة ويستبدلها بفراغ - وده
    # اللي كان بيسبب اختفاء العروض أو ظهور "-" بدل اسم المنتج بعد أي
    # rescan جزئي الفشل.
    offers_asins_in_batch = {o.get("asin") for o in new_offer_records if o.get("asin")}
    product_asins_in_batch = set()
    if product_record and product_record.get("asin"):
        product_asins_in_batch.add(product_record["asin"])

    existing_offer_rows = _read_existing_rows(path, "Offers")
    existing_product_rows = _read_existing_rows(path, "Products")

    kept_offer_rows = [r for r in existing_offer_rows if r[0] not in offers_asins_in_batch]
    kept_product_rows = [r for r in existing_product_rows if r[0] not in product_asins_in_batch]

    # لو الـ rescan رجع بيانات منتج ناقصة (مثلاً فشل يجيب العنوان/الوصف/
    # الصورة) بس كان فيه صف قديم لنفس الـ ASIN بقيم حقيقية، نحافظ على
    # القيم القديمة بدل ما نمسحها بـ "-"
    old_product_by_asin = {
        row[0]: row for row in existing_product_rows
        if row[0] in product_asins_in_batch
    }

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()

    # -------- Offers --------
    offers_ws = wb.active
    offers_ws.title = "Offers"
    _style_header(offers_ws, OFFERS_HEADERS, header_fill, header_font, border)
    widths = {"A": 14, "B": 45, "C": 18, "D": 16, "E": 30,
              "F": 14, "G": 14, "H": 15, "I": 12, "J": 12, "K": 14}
    for col, width in widths.items():
        offers_ws.column_dimensions[col].width = width
    offers_ws.freeze_panes = "A2"
    offers_ws.sheet_view.rightToLeft = True

    def write_offer_row(row_idx, values, seller):
        for col, value in enumerate(values, 1):
            cell = offers_ws.cell(row_idx, col, value if value is not None else "-")
            cell.border = border
            if col in [6, 7, 8, 9, 10]:
                cell.number_format = '0.00 "SAR"'
            if col == 11 and isinstance(value, (int, float)):
                cell.number_format = '0.0'
            if col == 2 and value and value != "-":
                cell.hyperlink = value
                cell.font = hyperlink_font
        is_my_offer = (
            my_seller_name and seller
            and my_seller_name.strip().lower() in str(seller).strip().lower()
        )
        if is_my_offer:
            for col in range(1, len(OFFERS_HEADERS) + 1):
                offers_ws.cell(row_idx, col).fill = my_offer_fill

    row_idx = 2
    for row in kept_offer_rows:
        write_offer_row(row_idx, list(row), row[4] if len(row) > 4 else None)
        row_idx += 1
    for offer in new_offer_records:
        values = [
            offer.get("asin"), offer.get("product_url"), offer.get("date"),
            offer.get("offer_id"), offer.get("seller"), offer.get("product_price"),
            offer.get("original_price"), offer.get("discount_amount"),
            offer.get("shipping"), offer.get("total"), offer.get("seller_rating"),
        ]
        write_offer_row(row_idx, values, offer.get("seller"))
        row_idx += 1

    if offers_ws.max_row >= 1:
        offers_ws.auto_filter.ref = f"A1:K{offers_ws.max_row}"

    # -------- Products --------
    products_ws = wb.create_sheet("Products")
    _style_header(products_ws, PRODUCTS_HEADERS, header_fill, header_font, border)
    widths = {"A": 14, "B": 45, "C": 70, "D": 60, "E": 45}
    for col, width in widths.items():
        products_ws.column_dimensions[col].width = width
    products_ws.freeze_panes = "A2"
    products_ws.sheet_view.rightToLeft = True

    def write_product_row(row_idx, values):
        for col, value in enumerate(values, 1):
            cell = products_ws.cell(row_idx, col, value if value is not None else "-")
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 3))
            if col in [4, 5] and value and value != "-":
                cell.hyperlink = value
                cell.font = hyperlink_font

    row_idx = 2
    for row in kept_product_rows:
        write_product_row(row_idx, list(row))
        row_idx += 1
    if product_record:
        old_row = old_product_by_asin.get(product_record.get("asin"))

        def _value_or_fallback(new_value, old_index):
            """لو القيمة الجديدة فاضية بس فيه قيمة قديمة حقيقية (مش '-')، استخدم القديمة."""
            if new_value:
                return new_value
            if old_row and len(old_row) > old_index and old_row[old_index] and old_row[old_index] != "-":
                return old_row[old_index]
            return new_value

        values = [
            product_record.get("asin"),
            _value_or_fallback(product_record.get("title"), 1),
            _value_or_fallback(product_record.get("description"), 2),
            _value_or_fallback(product_record.get("image_url"), 3),
            _value_or_fallback(product_record.get("product_url"), 4),
        ]
        write_product_row(row_idx, values)
        row_idx += 1

    if products_ws.max_row >= 1:
        products_ws.auto_filter.ref = f"A1:E{products_ws.max_row}"

    try:
        wb.save(path)
        print(f"[+] Last-data file updated: {path} "
              f"(منتجات: {products_ws.max_row - 1}, عروض: {offers_ws.max_row - 1})")
    except PermissionError:
        print(f"[!] الملف {path} مفتوح في برنامج تاني (زي Excel) — اقفله وجرب تاني.")
        raise
